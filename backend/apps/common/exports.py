"""CSV / Excel exports of the main business entities."""
import csv
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView

from apps.clients.models import Client
from apps.lots.models import Lot
from apps.transactions.models import Payment, Sale

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    XLSX = True
except Exception:
    XLSX = False


def _csv_response(filename: str, rows: list[dict]) -> HttpResponse:
    if not rows:
        rows = []
    buf = StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()), extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
    response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    ts = timezone.now().strftime("%Y%m%d-%H%M")
    response["Content-Disposition"] = f'attachment; filename="{filename}-{ts}.csv"'
    return response


def _xlsx_response(filename: str, rows: list[dict], sheet_title="Export") -> HttpResponse:
    if not XLSX:
        return _csv_response(filename, rows)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel limit
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        # Header styling: bold white on brand-blue
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF1D4ED8")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Rows
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        # Auto-width approx
        for col_idx, h in enumerate(headers, start=1):
            length = max(
                [len(str(h))] + [len(str(r.get(h, ""))) for r in rows[:200]]
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(length + 2, 50)
        ws.freeze_panes = "A2"
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    ts = timezone.now().strftime("%Y%m%d-%H%M")
    response["Content-Disposition"] = f'attachment; filename="{filename}-{ts}.xlsx"'
    return response


def _pick_format(request, filename, rows, sheet="Export"):
    # Use a custom param: "format" is intercepted by DRF content negotiation.
    fmt = (request.query_params.get("type")
           or request.query_params.get("export_format")
           or "csv").lower()
    if fmt in ("xlsx", "excel"):
        return _xlsx_response(filename, rows, sheet_title=sheet)
    return _csv_response(filename, rows)


class LotsExport(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = [{
            "Référence": l.reference,
            "Désignation": l.title,
            "Ville": l.city.name if l.city_id else "",
            "Quartier": l.neighborhood.name if l.neighborhood_id else "",
            "Type": l.get_lot_type_display(),
            "Surface (m²)": l.surface_m2,
            "Prix affiché": l.asking_price,
            "Devise": l.currency,
            "Statut": l.get_status_display(),
            "Eau": "Oui" if l.has_water else "Non",
            "Électricité": "Oui" if l.has_electricity else "Non",
            "Voirie": "Oui" if l.has_road_access else "Non",
            "Titre foncier": "Oui" if l.has_title_deed else "Non",
            "Référence cadastrale": l.cadastral_ref,
            "Date d'enregistrement": l.created_at.strftime("%d/%m/%Y") if l.created_at else "",
        } for l in Lot.objects.select_related("city", "neighborhood").all()]
        return _pick_format(request, "urban-land-lots", rows, "Lots")


class ClientsExport(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = [{
            "Code": c.code,
            "Type": c.get_kind_display(),
            "Nom / Raison sociale": c.display_name,
            "Email": c.email,
            "Téléphone": str(c.phone) if c.phone else "",
            "Pièce d'identité": c.get_id_kind_display() if c.id_kind else "",
            "N° de pièce": c.id_number,
            "Nationalité": c.nationality,
            "Adresse": c.address,
            "Ville": c.city,
            "Profession": c.profession,
            "Date d'enregistrement": c.created_at.strftime("%d/%m/%Y") if c.created_at else "",
            "Nombre de ventes": c.sales.count(),
        } for c in Client.objects.all()]
        return _pick_format(request, "urban-land-clients", rows, "Clients")


class SalesExport(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = [{
            "Référence": s.reference,
            "Date": s.sold_on.strftime("%d/%m/%Y") if s.sold_on else "",
            "Lot": s.lot.reference if s.lot_id else "",
            "Client": s.client.display_name if s.client_id else "",
            "Agent": s.agent.get_full_name() if s.agent_id else "",
            "Mode de paiement": s.get_payment_mode_display(),
            "Prix": s.price, "Remise": s.discount, "Net à payer": s.net_amount,
            "Encaissé": s.total_paid, "Solde dû": s.balance_due,
            "Avancement (%)": round(s.progress_pct, 1),
            "Statut": s.get_status_display(),
            "Apport initial": s.down_payment,
            "Nombre d'échéances": s.installment_count,
            "Fréquence (jours)": s.installment_frequency_days,
            "Devise": s.currency,
        } for s in Sale.objects.select_related("lot", "client", "agent").prefetch_related("payments").all()]
        return _pick_format(request, "urban-land-ventes", rows, "Ventes")


class PaymentsExport(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        rows = [{
            "Reçu": p.receipt_number,
            "Date": p.paid_on.strftime("%d/%m/%Y") if p.paid_on else "",
            "Vente": p.sale.reference if p.sale_id else "",
            "Client": p.sale.client.display_name if p.sale_id and p.sale.client_id else "",
            "Montant": p.amount, "Devise": p.currency,
            "Mode": p.get_method_display(),
            "Référence transaction": p.reference,
            "Annulé": "Oui" if p.is_void else "Non",
            "Reçu par": p.received_by.get_full_name() if p.received_by_id else "",
            "Empreinte intégrité": (p.integrity_hash or "")[:16] + "…",
        } for p in Payment.objects.select_related("sale__client", "received_by").all()]
        return _pick_format(request, "urban-land-versements", rows, "Versements")
